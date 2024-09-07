import random
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import sys


# 设置递归深度
sys.setrecursionlimit(1000000)


# 退火算法
class SimulatedAnnealing:

    # 初始化
    def __init__(self, T, cooling_rate, max_iter, obj_func, list_forbidden, bean_flag, min_T, year, alpha):
        self.T = T    # 初始温度
        self.cooling_rate = cooling_rate    # 退火速率
        self.max_iter = max_iter    # 最大迭代次数
        self.min_T = min_T    # 最小温度
        self.obj_func = obj_func    # 目标函数
        self.record = []    # 记录每一步的目标函数值
        self.list_forbidden = list_forbidden    # 禁忌表
        self.bean_flag = bean_flag    # 各个地块是否种植豆类（1表示一定要种植豆类，0表示不一定要种植豆类）
        self.year = year    # 第几年
        self.alpha = alpha    # 投资偏好系数

    # 迭代主体
    def run(self):
        # 初始化参数
        initial_solution = self.generate_new_solution(np.zeros(shape=(59, 54)), 1)    # 随机生成初始解
        current_solution = initial_solution    # 将初始解赋给当前解
        best_solution = current_solution    # 记录最优解
        best_obj_value = self.obj_func(best_solution, self.year, self.alpha)    # 记录最优解的目标函数值

        # 迭代
        for i in range(self.max_iter):
            new_solution = self.generate_new_solution(current_solution, 0.5)    # 生成新解
            new_obj_value = self.obj_func(new_solution, self.year, self.alpha)    # 计算新解的目标函数值
            delta_obj = new_obj_value - best_obj_value    # 计算新解与最优解的目标函数差值
            # 接受或拒绝新解
            if delta_obj > 0 or (delta_obj <= 0 and random.random() < np.exp(delta_obj / self.T)):
                # 接受新解
                current_solution = new_solution
                # 更新最优解
                if new_obj_value > best_obj_value:
                    best_solution = new_solution
                    best_obj_value = new_obj_value
            # 更新温度
            self.T *= self.cooling_rate
            if self.T < self.min_T:
                break
            # 记录每一步的目标函数值
            self.record.append((i, self.obj_func(current_solution, self.year, self.alpha)))
            # 打印当前状态
            print(self.record[-1])
        return best_solution

    # 约束条件测试
    def test(self, current_solution):
        # 每种作物同一季不能分散在超过7个地块上种植
        # 遍历每一行的非零元素个数
        for row in np.count_nonzero(current_solution, axis=1):
            if row > 7:
                return False
        return True

    # 生成新解
    def generate_new_solution(self, current_solution, change_chance):    # change_chance表示随机变异的概率
        new_solution = np.zeros(shape=np.shape(current_solution))    # 初始化新解
        # 判断水浇地种植季次策略
        flag = np.zeros(shape=(8))    # 记录水浇地种植季次策略
        # 遍历8种水浇地
        for i in range(8):
            if self.bean_flag[26 + i] == 1:    # 如果该地块一定要种植豆类，则必须双季种植
                flag[i] = 0
            elif [15, i + 26] in self.list_forbidden:    # 如果去年该地块种植过水稻，则必须双季种植
                flag[i] = 0
            else:
                # 随机选择水浇地种植季次策略
                if random.random() <= 0.5:
                    flag[i] = 1    # 1表示单季种植，0表示双季种植
                else:
                    flag[i] = 0
        # 遍历54个地块
        for j in range(54):
            # 随机判断该地块种植情况是否改变
            if random.random() <= change_chance:
                # 按地块类型分类讨论
                if 0 <= j < 26:    # 如果是平旱地、梯田、山坡地
                    if self.bean_flag[j] == 1:    # 如果该地块一定要种植豆类
                        # 随机选择一个粮食（豆类）
                        rand = random.randint(0, 4)
                        while [rand, j] in self.list_forbidden:    # 如果该粮食（豆类）在禁忌表中，则重新选择
                            rand = random.randint(0, 4)
                        new_solution[rand, j] += 0.5
                        # 随机选择另一个粮食（除水稻）
                        rand = random.randint(0, 14)
                        while [rand, j] in self.list_forbidden:    # 如果该粮食在禁忌表中，则重新选择
                            rand = random.randint(0, 14)
                        new_solution[rand, j] += 0.5
                    else:    # 如果该地块不一定要种植豆类
                        # 随机选择两种粮食（除水稻）
                        for k in range(2):
                            rand = random.randint(0, 14)
                            while [rand, j] in self.list_forbidden:
                                rand = random.randint(0, 14)
                            new_solution[rand, j] += 0.5
                elif 26 <= j < 34:    # 如果是水浇地
                    if flag[j - 26] == 1:    # 如果该地块单季种植
                        new_solution[15, j] += 1    # 只能全部种植水稻
                    elif flag[j - 26] == 0:     # 如果该地块双季种植
                        if self.bean_flag[j] == 1:     # 如果该地块一定要种植豆类
                            # 随机选择一种蔬菜（豆类）（第一季）
                            rand = random.randint(0, 2)
                            while [rand * 2 + 16, j] in self.list_forbidden:
                                rand = random.randint(0, 2)
                            new_solution[rand * 2 + 16, j] += 0.5
                            # 随机选择另一种蔬菜（第一季）
                            rand = random.randint(0, 17)
                            while [rand * 2 + 16, j] in self.list_forbidden:
                                rand = random.randint(0, 17)
                            new_solution[rand * 2 + 16, j] += 0.5
                            # 随机选择两种蔬菜（水浇地第二季）
                            for k in range(2):
                                rand = random.randint(0, 2)
                                while [rand + 52, j] in self.list_forbidden:
                                    rand = random.randint(0, 2)
                                new_solution[rand + 52, j] += 0.5
                        else:
                            # 随机选择两种蔬菜（第一季）
                            for k in range(2):
                                rand = random.randint(0, 17)
                                while [rand * 2 + 16, j] in self.list_forbidden:
                                    rand = random.randint(0, 17)
                                new_solution[rand * 2 + 16, j] += 0.5
                            # 随机选择两种蔬菜（水浇地第二季）
                            for k in range(2):
                                rand = random.randint(0, 2)
                                while [rand + 52, j] in self.list_forbidden:
                                    rand = random.randint(0, 2)
                                new_solution[rand + 52, j] += 0.5
                elif 34 <= j < 50:    # 如果是普通大棚
                    if self.bean_flag[j] == 1:    # 如果该地块一定要种植豆类
                        # 随机选择一种蔬菜（豆类）（第一季）
                        rand = random.randint(0, 2)
                        while [rand * 2 + 16, j] in self.list_forbidden:
                            rand = random.randint(0, 2)
                        new_solution[rand * 2 + 16, j] += 0.5
                        # 随机选择另一种蔬菜（第一季）
                        rand = random.randint(0, 17)
                        while [rand * 2 + 16, j] in self.list_forbidden:
                            rand = random.randint(0, 17)
                        new_solution[rand * 2 + 16, j] += 0.5
                        # 随机选择两种食用菌
                        for k in range(2):
                            rand = random.randint(0, 2)
                            while [rand + 55, j] in self.list_forbidden:
                                rand = random.randint(0, 2)
                            new_solution[rand + 55, j] += 0.5
                    else:
                        # 随机选择两种蔬菜（第一季）
                        for k in range(2):
                            rand = random.randint(0, 17)
                            while [rand * 2 + 16, j] in self.list_forbidden:
                                rand = random.randint(0, 17)
                            new_solution[rand * 2 + 16, j] += 0.5
                        # 随机选择两种食用菌
                        for k in range(2):
                            rand = random.randint(0, 2)
                            while [rand + 55, j] in self.list_forbidden:
                                rand = random.randint(0, 2)
                            new_solution[rand + 55, j] += 0.5
                else:    # 如果是智慧大棚
                    if self.bean_flag[j] == 1:    # 如果该地块一定要种植豆类
                        # 随机选择第一季还是第二季种豆类
                        if random.random() <= 0.5:
                            # 随机选择一种蔬菜（豆类）（第一季）
                            rand = random.randint(0, 2)
                            while [rand * 2 + 16, j] in self.list_forbidden:
                                rand = random.randint(0, 2)
                            new_solution[rand * 2 + 16, j] += 0.5
                            # 随机选择另一种蔬菜（第一季）
                            rand = random.randint(0, 17)
                            while [rand * 2 + 16, j] in self.list_forbidden:
                                rand = random.randint(0, 17)
                            new_solution[rand * 2 + 16, j] += 0.5
                            # 随机选择两种蔬菜（智慧大棚第二季）
                            for k in range(2):
                                rand = random.randint(0, 17)
                                while [rand * 2 + 16 + 1, j] in self.list_forbidden:
                                    rand = random.randint(0, 17)
                                new_solution[rand * 2 + 16 + 1, j] += 0.5
                        else:
                            # 随机选择两种蔬菜（第一季）
                            for k in range(2):
                                rand = random.randint(0, 17)
                                while [rand * 2 + 16, j] in self.list_forbidden:
                                    rand = random.randint(0, 17)
                                new_solution[rand * 2 + 16, j] += 0.5
                            # 随机选择一种蔬菜（豆类）（第二季）
                            rand = random.randint(0, 2)
                            while [rand * 2 + 16 + 1, j] in self.list_forbidden:
                                rand = random.randint(0, 2)
                            new_solution[rand * 2 + 16 + 1, j] += 0.5
                            # 随机选择另一种蔬菜（第二季）
                            rand = random.randint(0, 17)
                            while [rand * 2 + 16 + 1, j] in self.list_forbidden:
                                rand = random.randint(0, 17)
                            new_solution[rand * 2 + 16 + 1, j] += 0.5
                    else:    # 如果该地块不一定要种植豆类
                        # 随机选择两种蔬菜（第一季）
                        for k in range(2):
                            rand = random.randint(0, 17)
                            while [rand * 2 + 16, j] in self.list_forbidden:
                                rand = random.randint(0, 17)
                            new_solution[rand * 2 + 16, j] += 0.5
                        # 随机选择两种蔬菜（第二季）
                        for k in range(2):
                            rand = random.randint(0, 17)
                            while [rand * 2 + 16 + 1, j] in self.list_forbidden:
                                rand = random.randint(0, 17)
                            new_solution[rand * 2 + 16 + 1, j] += 0.5
            else:    # 如果该地块种植情况不改变
                new_solution[:, j] = current_solution[:, j]    # 直接复制当前解
        # 判断新解是否满足约束条件
        if not self.test(new_solution):
            return self.generate_new_solution(current_solution, change_chance)    # 若新解不满足约束条件，则重新生成
        else:
            return new_solution
    
    # 获取迭代过程
    def record(self):
        return self.record


if __name__ == '__main__':
    # 读取数据
    X = np.genfromtxt('planting_2023_matrix.csv', delimiter=',')
    C = np.genfromtxt('planting_cost_matrix.csv', delimiter=',')
    S = np.genfromtxt('planting_salesprice_matrix.csv', delimiter=',')
    Y = np.genfromtxt('planting_yield_matrix.csv', delimiter=',')
    A = np.genfromtxt('planting_area_matrix.csv', delimiter=',')

    # 定义目标函数
    def obj_func(solution, year, alpha):
        all = (Y * solution).sum(axis=1)    # 每种粮食总生产量
        anticipate = (Y * X).sum(axis=1)    # 每种粮食预期销售量

        anticipate_best = np.zeros(shape=(59))    # 每种粮食最优预期销售量
        for i in range(59):
            if i == 5 or i == 6:
                anticipate_best[i] = anticipate[i] * (1.1 ** (year + 1))
            else:
                anticipate_best[i] = anticipate[i] * (1.05 ** (year + 1))
        
        anticipate_worst = np.zeros(shape=(59))    # 每种粮食最差预期销售量
        for i in range(59):
            if i == 5 or i == 6:
                anticipate_worst[i] = anticipate[i] * (1.05 ** (year + 1))
            else:
                anticipate_worst[i] = anticipate[i] * (0.95 ** (year + 1))

        normal_best = np.minimum(all, anticipate_best)    # 每种粮食正常出售的部分
        normal_worst = np.minimum(all, anticipate_worst)    # 每种粮食正常出售的部分

        price_best = np.zeros(shape=(59))
        for i in range(59):
            if 16 <= i < 55:
                price_best[i] = S[i] * (1.05 ** (year + 1))
            elif 55 <= i < 58:
                price_best[i] = S[i] * (0.99 ** (year + 1))
            elif i == 58:
                price_best[i] = S[i] * (0.95 ** (year + 1))
            else:
                price_best[i] = S[i]
        
        price_worst = np.zeros(shape=(59))
        for i in range(59):
            if 16 <= i < 55:
                price_worst[i] = S[i] * (1.05 ** (year + 1))
            elif 55 <= i < 59:
                price_worst[i] = S[i] * (0.95 ** (year + 1))
            else:
                price_worst[i] = S[i]

        cost = (C * solution).sum() * (1.05 ** (year + 1))    # 总成本

        income_best = ((price_best * normal_best + (normal_best - anticipate_best) * 0.5) * price_best).sum()    # 最佳收益
        income_worst = ((price_worst * normal_worst + (normal_worst - anticipate_worst) * 0.5) * price_worst).sum()    # 最差收益

        risk_free_income = income_best - cost    # 无风险收益
        risk_premium = income_best - income_worst    # 风险溢价
        objective = (1 - alpha) * risk_free_income / (risk_premium + risk_free_income) + alpha * risk_premium / (risk_premium + risk_free_income)    # 目标函数

        return objective

    # 禁忌表
    list_forbidden = [[i, j]for [i, j] in np.argwhere(X > 0)]    # 初始禁忌表为2023年种植情况
    
    # 贪心最优解
    best_solution = np.zeros(shape=(7, 59, 54))

    # 是否需要种植豆类标志
    bean_flag = np.zeros(shape=(54))

    # 记录最优解函数值
    record = []

    # 投资者偏好系数
    alpha = 0.49
    # 实例化
    for i in range(7):

        # 判断是否种植豆类
        if i == 1:    # 如果是2025年， 需要对照23年和24年种植情况
            bean_flag = np.ones(shape=(54))    # 先假设全部都需要种植豆类
            for j in range(54):    # 遍历54个地块
                if 0 <= j < 26:    # 如果是平旱地、梯田、山坡地
                    if any(X[0:5, j] > 0) or any(best_solution[0, 0:5, j] > 0):    # 如果2023年或2024年该地块种植过粮食（豆类）
                        bean_flag[j] = 0    # 则今年不必种植豆类
                else:    # 如果是水浇地、普通大棚、智慧大棚
                    if any(X[16:22, j] > 0) or any(best_solution[0, 16:22, j] > 0):    # 如果2023年或2024年该地块种植过蔬菜（豆类）
                        bean_flag[j] = 0    # 则今年不必种植豆类
        elif i > 1:    # 如果是2026年以后， 只对照前两年种植情况
            bean_flag = np.ones(shape=(54))
            for j in range(X.shape[1]):
                if 0 <= j < 26:
                    if any(best_solution[i-1, 0:5, j] > 0) or any(best_solution[i-2, 0:5, j] > 0):
                        bean_flag[j] = 0
                else:
                    if any(best_solution[i-1, 16:22, j] > 0) or any(best_solution[i-2, 16:22, j] > 0):
                        bean_flag[j] = 0

        # 实例化SA算法
        sa = SimulatedAnnealing(T=300, cooling_rate=0.99, max_iter=10000, min_T=1e-6, obj_func=obj_func, list_forbidden=list_forbidden, bean_flag=bean_flag, year=i, alpha=alpha)
        best_solution[i] = sa.run()

        # 更新禁忌表
        list_forbidden = [[i, j] for [i, j] in np.argwhere((best_solution[i] == 1) | (best_solution[i] == 2))]    # 新禁忌表为当前解种植情况

        # 记录最优解函数值
        best_func_value = 0
        for i in range(7):
            best_func_value += obj_func(best_solution[i], i, alpha)
        record.append(best_func_value)

    # 随机生成随机变量数据
    point = np.zeros(shape=10000)    # 随机变量数据
    for i in range(10000):
        anticipate = (Y * X).sum(axis=1)    # 每种粮食预期销售量
        price = S
        cost = C
        for j in range(7):
            all = (Y * best_solution[j]).sum(axis=1)    # 每种粮食总生产量
            
            anticipate_ = np.zeros(shape=(59))    # 每种粮食最优预期销售量
            for k in range(59):
                if k == 5 or k == 6:
                    anticipate_[k] = anticipate[k] * np.random.uniform(1.05, 1.1)
                else:
                    anticipate_[k] = anticipate[k] * np.random.uniform(0.95, 1.05)
            
            anticipate = anticipate_

            normal = np.minimum(all, anticipate)    # 每种粮食正常出售的部分
            
            price_ = np.zeros(shape=(59))
            for k in range(59):
                if 16 <= k < 55:
                    price_[k] = price[k] * 1.05
                elif 55 <= i < 58:
                    price_[k] = price[k] * np.random.uniform(0.95, 0.99)
                elif i == 58:
                    price_[k] = price[k] * 0.95
                else:
                    price_[k] = price[k]
            price = price_

            cost = cost * np.random.uniform(1, 1.05)    # 总成本
            income = ((normal + (all - normal) * 0.5) * price).sum()    # 收入
            profit = income - (cost * best_solution[j]).sum()    # 利润
            point[i] += profit
    print(np.mean(point))    # 输出平均利润
    print(np.std(point))    # 输出标准差
    print(np.max(point))    # 输出最大利润
    print(np.min(point))    # 输出最小利润
    print(np.median(point))    # 输出中位数
    

    # 保存最优解
    excel_path = '.\\result2.xlsx'    # 保存路径
    wb = load_workbook(excel_path)    # 打开excel文件
    # 写入数据
    for i in range(7):
        sheet_mane = '20{}'.format(i+24)    # 表名为2024年至2029年
        sheet = wb[sheet_mane]    # 选择表
        typical_matrix = np.zeros(shape=(82, 41))    # 模板标准种植矩阵
        best_solution_T = best_solution[i].T    # 对最优解转置
        # 填充种植矩阵
        for j in range(82):
            if 0 <= j < 26:    # 如果是第一季的平旱地、梯田、山坡地（单季，即第一季）
                for k in range(15):    # 遍历15种粮食
                    typical_matrix[j, k] = best_solution_T[j, k] * A[j]
            elif 26 <= j < 34:    # 如果是第一季的水浇地
                typical_matrix[j, 15] = best_solution_T[j, 15] * A[j]    # 水稻
                for k in range(18):
                    typical_matrix[j, k + 16] = best_solution_T[j, k * 2 + 16] * A[j]
            elif 34 <= j < 50:
                for k in range(18):
                    typical_matrix[j, k + 16] = best_solution_T[j, k * 2 + 16] * A[j]
            elif 50 <= j < 54:
                for k in range(18):
                    typical_matrix[j, k + 16] = best_solution_T[j, k * 2 + 16] * A[j]
            elif 54 <= j < 62:
                typical_matrix[j, 15] = best_solution_T[j - 54 + 26, 15] * A[j - 54 + 26]
                for k in range(3):
                    typical_matrix[j, k + 34] = best_solution_T[j - 54 + 26, k + 52] * A[j - 54 + 26]
            elif 62 <= j < 78:
                for k in range(4):
                    typical_matrix[j, k + 37] = best_solution_T[j - 54 + 26, k + 55] * A[j - 54 + 26]
            else:
                for k in range(18):
                    typical_matrix[j, k + 16] = best_solution_T[j - 54 + 26, k * 2 + 17] * A[j - 54 + 26]
        for j in range(82):
            for k in range(41):
                sheet.cell(row=j+2, column=k+3, value=typical_matrix[j, k])
    wb.save(excel_path)